#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate acts and invoices from tanker-wash work log.

Default mode is DRY-RUN (no file changes).
Use --apply to generate *copies* of source files and write results there.

Requirements:
  pip install openpyxl

Notes:
- Preserves macros in XLSM via keep_vba=True.
- Matches customers by name (case-insensitive) using sheet "Заказчики Палладиум", column B.
- Excludes:
    * Customer names: "Палладиум-арт", "Физлица" (case-insensitive)
    * Rows with yellow or orange fill (any cell in the row)
- Marks processed orders by coloring ONLY the "Номер заказ-наряда" cell yellow.

Author: generated with ChatGPT
"""

from __future__ import annotations
from copy import copy
import re

import argparse
import calendar
import dataclasses
import datetime as dt
import math
import os

from pathlib import Path
from dotenv import load_dotenv

import re
import shutil
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from openpyxl.utils.cell import range_boundaries


# =========================
# PATH CONSTANTS (EDIT ME)
# =========================

load_dotenv()

BASE_PATH = os.getenv("BASE_PATH")

if not BASE_PATH:
    raise RuntimeError("Переменная окружения BASE_PATH не задана в .env")

BASE_PATH = Path(BASE_PATH)

if not BASE_PATH.exists():
    raise RuntimeError(f"BASE_PATH не существует: {BASE_PATH}")


ORDERS_XLSX = BASE_PATH / "2026" / "02_2026.xlsx"
ACTS_XLSM = BASE_PATH / "2026" / "02_2026_акты_счета.xlsm"
INVOICE_JOURNAL_XLSX = BASE_PATH / "журнал счетов.xlsx"
CLIENTS_XLSX = BASE_PATH / "Заказчики, Договора" / "Заказчики мойки.xlsx"

# Sheets
ORDERS_SHEET = "заказ-наряды"
CLIENTS_SHEET = "Заказчики Палладиум"
JOURNAL_SHEET = "счета"

# VAT
VAT_RATE = 0.20
EPS = 0.01

# Exclusions
SKIP_CLIENTS_NORMALIZED = {  # compare after normalize_name()
    "палладиум-арт",
    "физлица",
}

# Excluded colors (RGB, as Excel stores them in openpyxl)
YELLOW_RGB = "FFFFFF00"
ORANGE_RGB = "FFFFC000"

PROCESSED_FILL = PatternFill(patternType="solid", fgColor=YELLOW_RGB)


# =========================
# Helpers
# =========================

def normalize_name(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def safe_float(x) -> float:
    if x is None or x == "":
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    # Sometimes numbers come as strings with commas
    try:
        return float(str(x).replace(" ", "").replace(",", "."))
    except Exception:
        return 0.0


def last_day_of_month(year: int, month: int) -> dt.date:
    last = calendar.monthrange(year, month)[1]
    return dt.date(year, month, last)


def ddmmyy(d: dt.date) -> str:
    return d.strftime("%d%m%y")


def make_copy_path(src: str, suffix: str) -> str:
    base, ext = os.path.splitext(src)
    return f"{base}_{suffix}{ext}"


def cell_fill_rgb(cell) -> Optional[str]:
    fill = getattr(cell, "fill", None)
    if not fill:
        return None

    if fill.patternType != "solid":
        return None

    color = getattr(fill, "fgColor", None)
    if not color:
        return None

    rgb = getattr(color, "rgb", None)
    if not rgb:
        return None

    # openpyxl может вернуть объект RGB вместо строки
    rgb_str = str(rgb)

    return rgb_str.upper()


def row_has_excluded_color(ws, row_idx: int, min_col: int, max_col: int) -> bool:
    for col in range(min_col, max_col + 1):
        rgb = cell_fill_rgb(ws.cell(row=row_idx, column=col))
        if rgb in (YELLOW_RGB, ORANGE_RGB):
            return True
    return False


@dataclasses.dataclass
class WorkLine:
    name: str
    date: dt.date
    price_wo_vat: float  # BYN (without VAT)


@dataclasses.dataclass
class OrderRow:
    row_idx: int
    date: dt.date
    order_no: str
    client_raw: str
    sum_total: float  # "Сумма" column (without VAT)
    currency: str

    # outside wash
    gov1: str
    type1: str
    price1: float
    gov2: str
    type2: str
    price2: float

    # internal wash
    sections: str
    total_volume: str
    product: str
    price3: float

    # steaming
    steam_hours_raw: object
    price5: float

    # extras
    extra_name: str
    extra_qty: object
    price6: float


def parse_orders(ws, year: int, month: int) -> Tuple[List[OrderRow], Dict[str, int]]:
    """
    Parse orders from the journal sheet.
    Returns list of OrderRow and a header->col_index map.
    """
    # Find header row: in your file it's row 5 (contains "Номер заказ-наряда")
    header_row = None
    header_map: Dict[str, int] = {}
    for r in range(1, 20):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(v == "Номер заказ-наряда" for v in row_vals):
            header_row = r
            for c, v in enumerate(row_vals, start=1):
                if isinstance(v, str) and v.strip():
                    header_map[v.strip()] = c
            break
    if not header_row:
        raise RuntimeError("Не найден заголовок таблицы в листе заказ-нарядов.")

    # Required columns
    required = [
        "Дата",
        "Номер заказ-наряда",
        "Заказчик",
        "Сумма",
        "Валюта",
        "Гос № (1)",
        "Тип (1)",
        "Цена (без НДС) (1)",
        "Гос № (2)",
        "Тип (2)",
        "Цена (без НДС) (2)",
        "Кол-во секций",
        "Общий объём",
        "Продукт",
        "Цена (без НДС) (3)",
        "Пар, ч",
        "Цена (без НДС) (5)",
        "Доп работы",
        "Количество",
        "Цена (без НДС) (6)",
    ]
    missing = [c for c in required if c not in header_map]
    if missing:
        raise RuntimeError(f"В заказ-нарядах не найдены колонки: {missing}")

    # Parse rows
    parsed: List[OrderRow] = []
    first_data_row = header_row + 1

    # Determine scan bounds for color checks
    min_col, max_col = 1, ws.max_column

    for r in range(first_data_row, ws.max_row + 1):
        date_val = ws.cell(r, header_map["Дата"]).value
        if not isinstance(date_val, (dt.datetime, dt.date)):
            continue
        d = date_val.date() if isinstance(date_val, dt.datetime) else date_val
        if d.year != year or d.month != month:
            continue

        client_raw = ws.cell(r, header_map["Заказчик"]).value
        if not client_raw:
            continue

        # Exclude by color (yellow/orange)
        if row_has_excluded_color(ws, r, min_col=min_col, max_col=max_col):
            continue

        order_no = ws.cell(r, header_map["Номер заказ-наряда"]).value
        if order_no is None:
            continue
        order_no = str(order_no).strip()

        currency = ws.cell(r, header_map["Валюта"]).value
        currency = (str(currency).strip() if currency else "BYN")

        sum_total = safe_float(ws.cell(r, header_map["Сумма"]).value)

        row = OrderRow(
            row_idx=r,
            date=d,
            order_no=order_no,
            client_raw=str(client_raw).strip(),
            sum_total=sum_total,
            currency=currency,

            gov1=str(ws.cell(r, header_map["Гос № (1)"]).value or "").strip(),
            type1=str(ws.cell(r, header_map["Тип (1)"]).value or "").strip(),
            price1=safe_float(ws.cell(r, header_map["Цена (без НДС) (1)"]).value),

            gov2=str(ws.cell(r, header_map["Гос № (2)"]).value or "").strip(),
            type2=str(ws.cell(r, header_map["Тип (2)"]).value or "").strip(),
            price2=safe_float(ws.cell(r, header_map["Цена (без НДС) (2)"]).value),

            sections=str(ws.cell(r, header_map["Кол-во секций"]).value or "").strip(),
            total_volume=str(ws.cell(r, header_map["Общий объём"]).value or "").strip(),
            product=str(ws.cell(r, header_map["Продукт"]).value or "").strip(),
            price3=safe_float(ws.cell(r, header_map["Цена (без НДС) (3)"]).value),

            steam_hours_raw=ws.cell(r, header_map["Пар, ч"]).value,
            price5=safe_float(ws.cell(r, header_map["Цена (без НДС) (5)"]).value),

            extra_name=str(ws.cell(r, header_map["Доп работы"]).value or "").strip(),
            extra_qty=ws.cell(r, header_map["Количество"]).value,
            price6=safe_float(ws.cell(r, header_map["Цена (без НДС) (6)"]).value),
        )
        parsed.append(row)

    return parsed, header_map


def load_clients_map(path: str) -> Dict[str, str]:
    """
    Returns normalized_name -> canonical_name (exact as in reference sheet column B).
    """
    wb = load_workbook(path, data_only=True)
    if CLIENTS_SHEET not in wb.sheetnames:
        raise RuntimeError(f"В справочнике нет листа '{CLIENTS_SHEET}'")
    ws = wb[CLIENTS_SHEET]
    m: Dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 2).value  # column B
        if not name:
            continue
        canon = str(name).strip()
        key = normalize_name(canon)
        if key:
            m[key] = canon
    return m


def parse_hours(value) -> Optional[float]:
    """
    Пар, ч can be:
      - number (hours)
      - time (datetime.time) meaning HH:MM duration
      - string like '1,5' or '1.5'
    """
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, dt.time):
        return value.hour + value.minute / 60.0 + value.second / 3600.0
    # sometimes Excel stores durations as datetime.datetime with 1899 base, but rarely here
    if isinstance(value, dt.datetime):
        t = value.time()
        return t.hour + t.minute / 60.0 + t.second / 3600.0
    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except Exception:
        return None


def build_work_lines(order: OrderRow) -> List[WorkLine]:
    """
    Build multiple lines per one order row based on your rules.
    Outside wash: ONE LINE combining (1) and (2).
    """
    lines: List[WorkLine] = []
    d = order.date

    # 1) Outside wash: if any type present
    outside_parts: List[str] = []
    outside_price = 0.0

    if order.type1:
        part = order.type1
        if order.gov1:
            part += f" {order.gov1}"
        outside_parts.append(part)
        outside_price += order.price1

    if order.type2:
        part = order.type2
        if order.gov2:
            part += f" {order.gov2}"
        outside_parts.append(part)
        outside_price += order.price2

    if outside_parts and outside_price > 0:
        name = "мойка: " + "; ".join(outside_parts)
        lines.append(WorkLine(name=name, date=d, price_wo_vat=outside_price))

    # 2) Internal wash
    if order.price3 > 0:
        gov = order.gov2 or order.gov1  # usually the tanker plate is (2)
        name = f"внутренняя мойка цистерны {gov}".strip()
        details: List[str] = []
        if order.sections:
            details.append(f"количество секций: {order.sections}")
        if order.total_volume:
            details.append(f"объем: {order.total_volume}")
        if order.product:
            details.append(f"продукт {order.product}")
        if details:
            name = name + ", " + ", ".join(details)
        lines.append(WorkLine(name=name, date=d, price_wo_vat=order.price3))

    # 3) Steam
    h = parse_hours(order.steam_hours_raw)
    if h is not None and h > 0 and order.price5 > 0:
        # pretty hours (avoid .0)
        if abs(h - round(h)) < 1e-9:
            hs = str(int(round(h)))
        else:
            hs = str(round(h, 2)).rstrip("0").rstrip(".")
        name = f"пропарка: {hs} ч"
        lines.append(WorkLine(name=name, date=d, price_wo_vat=order.price5))

    # 4) Extra works
    if order.extra_name and order.price6 > 0:
        qty = order.extra_qty
        qty_s = ""
        if qty is not None and str(qty).strip() != "":
            # keep integer if possible
            qf = safe_float(qty)
            if abs(qf - round(qf)) < 1e-9:
                qty_s = f", {int(round(qf))} шт"
            else:
                qty_s = f", {str(qf).rstrip('0').rstrip('.')} шт"
        name = f"{order.extra_name}{qty_s}"
        lines.append(WorkLine(name=name, date=d, price_wo_vat=order.price6))

    return lines


def group_by_client(orders: List[OrderRow], clients_map: Dict[str, str]) -> Tuple[Dict[str, List[OrderRow]], List[str]]:
    grouped: Dict[str, List[OrderRow]] = {}
    not_matched: List[str] = []

    for o in orders:
        key = normalize_name(o.client_raw)
        if key in SKIP_CLIENTS_NORMALIZED:
            continue

        canon = clients_map.get(key)
        if not canon:
            not_matched.append(o.client_raw)
            continue

        grouped.setdefault(canon, []).append(o)

    # de-duplicate not_matched preserving order
    seen = set()
    uniq = []
    for x in not_matched:
        k = normalize_name(x)
        if k not in seen:
            seen.add(k)
            uniq.append(x)
    return grouped, uniq


def compute_sums_for_client(orders: List[OrderRow]) -> Tuple[float, float, int]:
    """
    Returns (sum_source, sum_built, work_lines_count)
    """
    sum_source = 0.0
    sum_built = 0.0
    lines_count = 0
    for o in orders:
        sum_source += o.sum_total
        lines = build_work_lines(o)
        lines_count += len(lines)
        sum_built += sum(l.price_wo_vat for l in lines)
    return sum_source, sum_built, lines_count


def find_next_doc_number(journal_path: str, date_suffix: str) -> int:
    """
    Reads invoice journal and returns next sequential integer for numbers ending with /<date_suffix>.
    e.g. existing: 001/280226 -> returns 2
    """
    wb = load_workbook(journal_path, data_only=True)
    if JOURNAL_SHEET not in wb.sheetnames:
        raise RuntimeError(f"В журнале счетов нет листа '{JOURNAL_SHEET}'")
    ws = wb[JOURNAL_SHEET]

    # Find column "номер счета"
    header_row = 1
    col_num = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and normalize_name(v) in ("номер счета", "номер счёта"):
            col_num = c
            break
    if not col_num:
        # fallback: column B per your file
        col_num = 2

    best = 0
    pattern = re.compile(r"^\s*(\d+)\s*/\s*" + re.escape(date_suffix) + r"\s*$")
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col_num).value
        if not v:
            continue
        m = pattern.match(str(v))
        if m:
            n = int(m.group(1))
            best = max(best, n)

    return best + 1


def fmt_doc_number(n: int, date_suffix: str) -> str:
    return f"{n:03d}/{date_suffix}"

 
def update_excel_table_ref(ws, table_name: str, new_last_row: int) -> None:
    """
    Expand an existing Excel Table to reach new_last_row.
    Keeps same columns as original table.
    """
    if not ws.tables or table_name not in ws.tables:
        raise RuntimeError(f"Не найдена таблица '{table_name}' на листе '{ws.title}'")
    t = ws.tables[table_name]
    # t.ref like "A15:G18"
    start_cell, end_cell = t.ref.split(":")
    start_col_letters = re.findall(r"[A-Z]+", start_cell)[0]
    start_row = int(re.findall(r"\d+", start_cell)[0])
    end_col_letters = re.findall(r"[A-Z]+", end_cell)[0]
    t.ref = f"{start_col_letters}{start_row}:{end_col_letters}{new_last_row}"


def clone_sheet_from_template(wb, template_name: str, new_name: str, *, suffix: str) -> tuple:
    ws_tpl = wb[template_name]
    ws_new = wb.copy_worksheet(ws_tpl)
    ws_new.title = new_name

    # Восстанавливаем таблицы и получаем mapping имён
    name_map = clone_tables_from_template(ws_tpl, ws_new, suffix=suffix)

    return ws_new, name_map

def replace_in_formulas(ws, old: str, new: str) -> None:
    """
    Заменяем structured references в формулах:
    =ТаблицаАкт[[#This Row],[Сумма]]  ->  =ТаблицаАкт_280226[[#This Row],[Сумма]]
    """
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("=") and old in cell.value:
                cell.value = cell.value.replace(old, new)


def clone_tables_from_template(template_ws, target_ws, suffix: str) -> dict[str, str]:
    """
    Воссоздаёт Excel Tables на target_ws (copy_worksheet их не переносит).
    КРИТИЧНО: создаём tableColumns по заголовкам, иначе Excel может не открыть файл.
    Возвращает mapping old_name -> new_name.
    """
    name_map: dict[str, str] = {}

    for old_name, _ref_str in template_ws.tables.items():
        old_table = template_ws.tables[old_name]  # Table object
        new_name = f"{old_name}_{suffix}"
        name_map[old_name] = new_name

        ref = old_table.ref  # например "A15:G16" (header + 1 строка)
        min_col, min_row, max_col, max_row = range_boundaries(ref)

        # Создаём таблицу
        new_table = Table(displayName=new_name, ref=ref)

        # Стиль таблицы
        sti = getattr(old_table, "tableStyleInfo", None)
        if sti:
            new_table.tableStyleInfo = TableStyleInfo(
                name=sti.name,
                showFirstColumn=sti.showFirstColumn,
                showLastColumn=sti.showLastColumn,
                showRowStripes=sti.showRowStripes,
                showColumnStripes=sti.showColumnStripes,
            )

        # Header/totals
        new_table.headerRowCount = getattr(old_table, "headerRowCount", 1)
        new_table.totalsRowCount = getattr(old_table, "totalsRowCount", 0)

        # >>> ВАЖНО: tableColumns (берём имена из строки заголовка на target_ws)
        header_row = min_row
        cols: list[TableColumn] = []
        for i, col in enumerate(range(min_col, max_col + 1), start=1):
            header_val = target_ws.cell(row=header_row, column=col).value
            header_name = str(header_val).strip() if header_val not in (None, "") else f"Column{i}"
            cols.append(TableColumn(id=i, name=header_name))

        new_table.tableColumns = cols
        # new_table.tableColumns.count = len(cols)

        # AutoFilter (желательно, чтобы Excel не ругался)
        try:
            new_table.autoFilter = old_table.autoFilter
        except Exception:
            pass

        target_ws.add_table(new_table)

        # Если ты используешь structured references (ТаблицаАкт[[#This Row],...]),
        # то дальше нужно заменить имя таблицы в формулах:
        replace_in_formulas(target_ws, old_name, new_name)

    return name_map

def write_act_sheet(ws_act, *, table_name: str, client_name: str, doc_no: str, doc_date: dt.date,
                    contract_no: Optional[str], contract_date: Optional[dt.date], work_lines: List[WorkLine]):
    """
    Fill a cloned 'акт' sheet.
    Template assumptions based on your file:
      - doc number at D4
      - doc date at F4
      - client name at C9
      - contract at C12, contract date string at F12 (kept as-is if empty)
      - table 'ТаблицаАкт' starts at row 15 header, first data row 16
    """

    if not ws_act.tables:
        raise RuntimeError("На листе акта нет Excel-таблиц (после клонирования).")

    if table_name not in ws_act.tables:
        raise RuntimeError(f"На листе акта нет таблицы '{table_name}'. Есть: {list(ws_act.tables.keys())}")

    t = ws_act.tables[table_name]
    
    ws_act["D4"].value = doc_no
    ws_act["F4"].value = dt.datetime.combine(doc_date, dt.time(0, 0))

    ws_act["C9"].value = client_name

    if contract_no:
        ws_act["C12"].value = contract_no
    if contract_date:
        # Your template stores contract date as a string in F12, like «19» января 2026
        # We'll write a standard dd.mm.yyyy string there (safe), unless you prefer Russian quotes later.
        ws_act["F12"].value = contract_date.strftime("%d.%m.%Y")

    # Table
    header_row = 15
    first_data_row = 16

    # Determine how many template rows already exist in table
    # We'll assume the current table ref defines the available rows.
    t = ws_act.tables[table_name]
    start_cell, end_cell = t.ref.split(":")
    end_row = int(re.findall(r"\d+", end_cell)[0])
    existing_rows = max(0, end_row - first_data_row + 1)

    needed = len(work_lines)
    if needed == 0:
        return

    # Insert rows if needed (before totals section; totals are below table, so inserting inside table range is ok)
    if needed > existing_rows:
        insert_at = first_data_row + existing_rows
        ws_act.insert_rows(insert_at, amount=(needed - existing_rows))

    # Write rows
    for i, wl in enumerate(work_lines, start=1):
        r = first_data_row + (i - 1)
        ws_act.cell(r, 1).value = i
        ws_act.cell(r, 2).value = wl.name
        ws_act.cell(r, 3).value = dt.datetime.combine(wl.date, dt.time(0, 0))
        ws_act.cell(r, 4).value = round(wl.price_wo_vat, 2)
        ws_act.cell(r, 5).value = VAT_RATE
        # Columns 6-7 in template have formulas referencing the table; we keep them
        # but make sure they exist
        if ws_act.cell(r, 6).value in (None, ""):
            ws_act.cell(r, 6).value = f"={table_name}[[#This Row],[Сумма]]*{table_name}[[#This Row],[Ставка НДС]]"
        if ws_act.cell(r, 7).value in (None, ""):
            ws_act.cell(r, 7).value = f"={table_name}[[#This Row],[Сумма]]+{table_name}[[#This Row],[Сумма НДС]]"

    new_last_row = first_data_row + needed - 1
    update_excel_table_ref(ws_act, table_name, new_last_row)


def write_invoice_sheet(ws_inv, *, table_name: str, client_info_text: str, doc_no: str, doc_date: dt.date, contract_no: Optional[str], contract_date: Optional[dt.date], work_lines: List[WorkLine]):
    """
    Fill a cloned 'счет' sheet.
    Template assumptions:
      - doc number at I2
      - doc date at J2
      - payer info at C5 (multiline)
      - contract no at C8, contract date at E8
      - table header at row 10, data starts row 11
      - VAT rate is constant
    """

    if not ws_inv.tables:
        raise RuntimeError("На листе акта нет Excel-таблиц (после клонирования).")

    if table_name not in ws_inv.tables:
        raise RuntimeError(f"На листе акта нет таблицы '{table_name}'. Есть: {list(ws_inv.tables.keys())}")

    t = ws_inv.tables[table_name]
    ws_inv["I2"].value = doc_no
    ws_inv["J2"].value = dt.datetime.combine(doc_date, dt.time(0, 0))

    ws_inv["C5"].value = client_info_text

    if contract_no:
        ws_inv["C8"].value = contract_no
    if contract_date:
        ws_inv["E8"].value = dt.datetime.combine(contract_date, dt.time(0, 0))

    header_row = 10
    first_data_row = 11
    table_name = None
    # find table if exists
    if ws_inv.tables:
        # pick first table
        table_name = list(ws_inv.tables.keys())[0]

    # Determine existing data rows in table
    if table_name:
        t = ws_inv.tables[table_name]
        start_cell, end_cell = t.ref.split(":")
        end_row = int(re.findall(r"\d+", end_cell)[0])
        existing_rows = max(0, end_row - first_data_row + 1)
    else:
        # fallback: assume 5 rows placeholder (11-15)
        existing_rows = 5

    needed = len(work_lines)
    if needed == 0:
        return

    if needed > existing_rows:
        insert_at = first_data_row + existing_rows
        ws_inv.insert_rows(insert_at, amount=(needed - existing_rows))

    for i, wl in enumerate(work_lines, start=1):
        r = first_data_row + (i - 1)
        ws_inv.cell(r, 1).value = i
        ws_inv.cell(r, 2).value = wl.name
        ws_inv.cell(r, 3).value = round(wl.price_wo_vat, 2)
        ws_inv.cell(r, 4).value = VAT_RATE
        # columns 5-6: VAT amount and with VAT. Keep formulas if exist, else set formulas
        if ws_inv.cell(r, 5).value in (None, ""):
            ws_inv.cell(r, 5).value = f"=C{r}*D{r}"
        if ws_inv.cell(r, 6).value in (None, ""):
            ws_inv.cell(r, 6).value = f"=C{r}+E{r}"

    new_last_row = first_data_row + needed - 1
    if table_name:
        update_excel_table_ref(ws_inv, table_name, new_last_row)


def build_client_info_for_invoice(client_name: str, clients_full_wb_path: str) -> Tuple[str, Optional[str], Optional[dt.date], Optional[str]]:
    """
    Build payer info text for invoice C5.
    Also returns (contract_no, contract_date, unp) if can be found.

    We try to find some typical headers in 'Заказчики Палладиум'. If not found, we fallback.
    """
    wb = load_workbook(clients_full_wb_path, data_only=True)
    ws = wb[CLIENTS_SHEET]

    # find row by name (case-insensitive)
    target_key = normalize_name(client_name)
    target_row = None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if v and normalize_name(str(v)) == target_key:
            target_row = r
            break
    if not target_row:
        # minimal
        return (client_name, None, None, None)

    # attempt to build header map from first row
    header_map: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        hv = ws.cell(1, c).value
        if isinstance(hv, str) and hv.strip():
            header_map[normalize_name(hv)] = c

    def get_by_header(*names: str) -> Optional[str]:
        for n in names:
            col = header_map.get(normalize_name(n))
            if col:
                v = ws.cell(target_row, col).value
                if v is None:
                    return None
                return str(v).strip()
        return None

    # Common possible headers
    unp = get_by_header("унп", "УНП")
    addr = get_by_header("адрес", "юридический адрес", "юр. адрес")
    contract_no = get_by_header("номер договора", "№ договора", "договор", "договор №")
    contract_date_raw = get_by_header("дата договора", "договор от", "дата")
    email = get_by_header("e-mail", "email", "почта")

    contract_date: Optional[dt.date] = None
    if contract_date_raw:
        # try parse dd.mm.yyyy
        try:
            contract_date = dt.datetime.strptime(contract_date_raw.replace("«", "").replace("»", "").strip(), "%d.%m.%Y").date()
        except Exception:
            contract_date = None

    # If the sheet stores contract date as excel date object
    if contract_date is None:
        # scan a few cols for a date if header matched
        for key in ("дата договора", "договор от"):
            col = header_map.get(key)
            if col:
                v = ws.cell(target_row, col).value
                if isinstance(v, (dt.datetime, dt.date)):
                    contract_date = v.date() if isinstance(v, dt.datetime) else v

    parts = [client_name]
    if unp:
        parts.append(f"УНП {unp}")
    if addr:
        parts.append(addr)
    info_text = "\n".join(parts)

    return info_text, contract_no, contract_date, unp


def append_to_invoice_journal(journal_path: str, doc_date: dt.date, doc_no: str, contract_no: Optional[str], contract_date: Optional[dt.date], client_name: str, unp: Optional[str], amount_wo_vat: float, currency: str, act_sheet_name: str, inv_sheet_name: str):
    """
    Append row into the invoice journal table on sheet 'счета'.
    We copy formulas for VAT and total from the previous data row.
    """
    wb = load_workbook(journal_path)
    ws = wb[JOURNAL_SHEET]

    # Find Excel table (first table on sheet)
    if not ws.tables:
        raise RuntimeError("На листе журнала счетов нет Excel-таблицы.")
    table_name = list(ws.tables.keys())[0]
    t = ws.tables[table_name]
    start_cell, end_cell = t.ref.split(":")
    start_row = int(re.findall(r"\d+", start_cell)[0])
    end_row = int(re.findall(r"\d+", end_cell)[0])

    # Find totals row: by your file, last row in table is totals row.
    totals_row = end_row

    # Insert a new row right before totals
    insert_at = totals_row
    ws.insert_rows(insert_at, amount=1)

    # Copy style & formulas from previous row (insert_at-1)
    src_row = insert_at - 1
    dst_row = insert_at

    max_col = ws.max_column
    for c in range(1, max_col + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        dst._style = copy(src._style)
        dst.number_format = src.number_format

        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
        dst.comment = src.comment  # если нужно, обычно не надо

        # copy formula/value
        if isinstance(src.value, str) and src.value.startswith("="):
            dst.value = src.value
        else:
            dst.value = None

    # Fill columns by header names (row 1)
    header = {normalize_name(ws.cell(1, c).value or ""): c for c in range(1, ws.max_column + 1)}

    def set_col(header_name_variants: List[str], value):
        for hn in header_name_variants:
            c = header.get(normalize_name(hn))
            if c:
                ws.cell(dst_row, c).value = value
                return

    set_col(["дата"], dt.datetime.combine(doc_date, dt.time(0, 0)))
    set_col(["номер счета", "номер счёта"], doc_no)
    set_col(["номер договора", "договор"], contract_no or "")
    set_col(["дата договора"], dt.datetime.combine(contract_date, dt.time(0, 0)) if contract_date else "")
    set_col(["наименование заказчика", "заказчик"], client_name)
    set_col(["унп"], unp or "")
    set_col(["стоимость без ндс"], round(amount_wo_vat, 2))
    set_col(["валюта"], currency)
    set_col(["акт"], act_sheet_name)
    set_col(["счет", "счёт"], inv_sheet_name)

    # Expand table ref by 1 row (totals row moved down)
    update_excel_table_ref(ws, table_name, new_last_row=totals_row + 1)

    wb.save(journal_path)
    wb.close()

def mark_processed_order(ws_orders, header_map: Dict[str, int], row_idx: int):
    col = header_map["Номер заказ-наряда"]
    cell = ws_orders.cell(row=row_idx, column=col)
    cell.fill = PROCESSED_FILL


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--year", type=int, required=True)
    p.add_argument("--month", type=int, required=True)
    p.add_argument("--apply", action="store_true", help="Apply changes into *_generated copies")
    args = p.parse_args()

    year, month = args.year, args.month

    period_last = last_day_of_month(year, month)
    suffix = ddmmyy(period_last)

    # Load clients map
    clients_map = load_clients_map(CLIENTS_XLSX)

    # Load orders (data_only True for values; but we also need colors so data_only doesn't matter)
    orders_wb = load_workbook(ORDERS_XLSX, data_only=True)
    if ORDERS_SHEET not in orders_wb.sheetnames:
        raise RuntimeError(f"В журнале заказ-нарядов нет листа '{ORDERS_SHEET}'")
    ws_orders = orders_wb[ORDERS_SHEET]

    orders, header_map = parse_orders(ws_orders, year, month)
    grouped, not_matched = group_by_client(orders, clients_map)

    # DRY-RUN report basics
    total_rows = len(orders)

    print("=== DRY RUN ===" if not args.apply else "=== APPLY (to copies) ===")
    print(f"Период: {month:02d}.{year} (номер с датой {suffix})")
    print(f"Строк заказ-нарядов (после фильтра по месяцу и по цветам): {total_rows}")

    if not_matched:
        print("\nНематченные заказчики (проверь названия в заказ-нарядах/справочнике):")
        for n in not_matched:
            print(f"  - {n}")

    # Build per-client summaries & validate sums
    client_summaries = []
    any_sum_errors = False

    for client_name, client_orders in sorted(grouped.items(), key=lambda kv: kv[0].lower()):
        sum_source, sum_built, lines_count = compute_sums_for_client(client_orders)
        diff = round(sum_built - sum_source, 2)
        ok = abs(diff) <= EPS
        if not ok:
            any_sum_errors = True
        client_summaries.append((client_name, len(client_orders), lines_count, sum_source, sum_built, diff, ok))

    print("\n--- Свод по заказчикам ---")
    for client_name, n_orders, n_lines, s_src, s_bld, diff, ok in client_summaries:
        status = "OK" if ok else "MISMATCH"
        print(f"{client_name}: заказ-нарядов={n_orders}, строк работ={n_lines}, сумма(таблица)={s_src:.2f}, сумма(акт)={s_bld:.2f}, Δ={diff:+.2f} [{status}]")

    if any_sum_errors:
        print("\nВНИМАНИЕ: обнаружены расхождения сумм. В режиме apply для таких заказчиков документы не будут созданы.")

    if not args.apply:
        print("\nФайлы НЕ изменены (dry-run).")
        return

    # APPLY: create copies
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    orders_out = make_copy_path(ORDERS_XLSX, f"generated_{ts}")
    acts_out = make_copy_path(ACTS_XLSM, f"generated_{ts}")
    journal_out = make_copy_path(INVOICE_JOURNAL_XLSX, f"generated_{ts}")

    shutil.copy2(ORDERS_XLSX, orders_out)
    shutil.copy2(ACTS_XLSM, acts_out)
    shutil.copy2(INVOICE_JOURNAL_XLSX, journal_out)

    print("\nСозданы копии:")
    print(f"  orders  -> {orders_out}")
    print(f"  acts    -> {acts_out}")
    print(f"  journal -> {journal_out}")

    # Reopen copies for writing
    orders_wb_w = load_workbook(orders_out)
    ws_orders_w = orders_wb_w[ORDERS_SHEET]

    acts_wb = load_workbook(acts_out, keep_vba=True)
    # journal will be updated per customer; we will open/save inside append function to keep table structure correct

    # Determine first doc number for the date suffix (from copied journal)
    next_n = find_next_doc_number(journal_out, suffix)

    created_docs = 0
    skipped_due_sum = 0

    for client_name, client_orders in sorted(grouped.items(), key=lambda kv: kv[0].lower()):
        sum_source, sum_built, lines_count = compute_sums_for_client(client_orders)
        diff = round(sum_built - sum_source, 2)
        if abs(diff) > EPS:
            print(f"SKIP (sum mismatch): {client_name} Δ={diff:+.2f}")
            skipped_due_sum += 1
            continue

        doc_no = fmt_doc_number(next_n, suffix)
        next_n += 1

        # Build work lines for this client, in order of date then order_no
        work_lines: List[WorkLine] = []
        # also collect rows to mark processed
        rows_to_mark: List[int] = []

        for o in sorted(client_orders, key=lambda x: (x.date, x.order_no, x.row_idx)):
            wls = build_work_lines(o)
            # If an order produced 0 lines but has sum_total > 0 -> warn
            if not wls and o.sum_total > 0:
                print(f"WARNING: заказ-наряд {o.order_no} ({client_name}) имеет сумму {o.sum_total} но не сформировал строк работ")
            work_lines.extend(wls)
            rows_to_mark.append(o.row_idx)

        client_info, contract_no, contract_date, unp = build_client_info_for_invoice(client_name, CLIENTS_XLSX)

        # Sheet names (sanitize)
        safe_client = re.sub(r"[^0-9A-Za-zА-Яа-я_ -]+", "", client_name)
        safe_client = safe_client.replace(" ", "_")
        act_sheet_name = f"акт_{safe_client}{doc_no.replace('/', '')}"
        inv_sheet_name = f"счет_{safe_client}{doc_no.replace('/', '')}"

        # Avoid collisions
        for nm in (act_sheet_name, inv_sheet_name):
            if nm in acts_wb.sheetnames:
                raise RuntimeError(f"Лист '{nm}' уже существует в xlsm. Остановлено чтобы не перезаписать архив.")

        tbl_suffix = doc_no.replace("/", "")

        ws_act, act_tables = clone_sheet_from_template(acts_wb, "акт", act_sheet_name, suffix=tbl_suffix)
        ws_inv, inv_tables = clone_sheet_from_template(acts_wb, "счет", inv_sheet_name, suffix=tbl_suffix)
        act_table_name = act_tables["ТаблицаАкт"]
        inv_table_name = inv_tables["ТаблицаСчет"]
        
        write_act_sheet(
            ws_act,
            table_name=act_table_name,
            client_name=client_name,
            doc_no=doc_no,
            doc_date=period_last,
            contract_no=contract_no,
            contract_date=contract_date,
            work_lines=work_lines,
        )
        write_invoice_sheet(
            ws_inv,
            table_name=inv_table_name,
            client_info_text=client_info,
            doc_no=doc_no,
            doc_date=period_last,
            contract_no=contract_no,
            contract_date=contract_date,
            work_lines=work_lines,
        )
        

        # Append to invoice journal
        append_to_invoice_journal(
            journal_path=journal_out,
            doc_date=period_last,
            doc_no=doc_no,
            contract_no=contract_no,
            contract_date=contract_date,
            client_name=client_name,
            unp=unp,
            amount_wo_vat=sum_built,
            currency="BYN",
            act_sheet_name=act_sheet_name,
            inv_sheet_name=inv_sheet_name,
        )

        # Mark processed orders: ONLY the "Номер заказ-наряда" cell
        for r in rows_to_mark:
            mark_processed_order(ws_orders_w, header_map, r)

        created_docs += 1
        print(f"CREATED: {client_name} -> {doc_no} (строк работ: {len(work_lines)}, сумма: {sum_built:.2f})")

    # Save updated workbooks
    acts_wb.save(acts_out)
    orders_wb_w.save(orders_out)
    acts_wb.close()
    orders_wb_w.close()

    print("\nГотово.")
    print(f"Создано комплектов акт+счет: {created_docs}")
    if skipped_due_sum:
        print(f"Пропущено из-за расхождения сумм: {skipped_due_sum}")
    print("Файлы результата (копии):")
    print(f"  {orders_out}")
    print(f"  {acts_out}")
    print(f"  {journal_out}")


if __name__ == "__main__":
    main()
